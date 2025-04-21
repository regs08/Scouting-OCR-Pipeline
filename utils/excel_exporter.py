import pandas as pd
from pathlib import Path
from typing import Union, Optional, List, Dict, Any
import logging
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import os

class ExcelExporter:
    """Utility class for exporting data to Excel format."""
    
    def __init__(self, output_dir: Optional[Union[str, Path]] = None):
        """
        Initialize the Excel exporter.
        
        Args:
            output_dir: Directory where Excel files will be saved
        """
        self.output_dir = Path(output_dir) if output_dir else Path.cwd() / "excel_exports"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.logger = logging.getLogger(__name__)
        
        # Define columns to flag
        self.columns_to_flag = [f'L{i}' for i in range(1, 21)]
        
        # Define values to flag
        self.values_to_flag = {10, 5, 1}
    
    def _flag_value(self, value: Any, column: str) -> str:
        """
        Flag values of 10, 5, and 1 in L1-L20 columns.
        
        Args:
            value: The value to potentially flag
            column: The column name this value is from
            
        Returns:
            The value, potentially flagged with ! if it's 10, 5, or 1 and in L1-L20
        """
        # Only flag if it's one of the L1-L20 columns
        if column not in self.columns_to_flag:
            return str(value)
            
        # Try to convert to integer if possible
        try:
            int_value = int(float(str(value).strip()))
            if int_value in self.values_to_flag:
                return f"{int_value}!"
        except (ValueError, TypeError):
            pass
            
        return str(value)

    def _convert_to_numeric(self, df: pd.DataFrame) -> pd.DataFrame:
        """Convert numeric columns to appropriate types."""
        for col in df.columns:
            # Try to convert to numeric, keeping integers as integers
            try:
                # First try to convert to integer
                int_series = pd.to_numeric(df[col], downcast='integer')
                if (int_series == df[col]).all():  # If all values match after conversion
                    df[col] = int_series
                else:
                    # If not all values match, try float
                    df[col] = pd.to_numeric(df[col], downcast='float')
            except (ValueError, TypeError):
                # If conversion fails, keep as string
                pass
        return df
    
    def export_csv_to_excel(
        self,
        csv_path: Union[str, Path],
        output_name: Optional[str] = None,
        sheet_name: str = "Data",
        index: bool = False,
        flag_values: Optional[List[Any]] = None  # This parameter is kept for backward compatibility but not used
    ) -> Path:
        """
        Convert a CSV file to Excel format, flagging values of 10, 5, and 1 in L1-L20 columns.
        
        Args:
            csv_path: Path to the CSV file
            output_name: Name for the output Excel file (without extension)
            sheet_name: Name of the sheet in the Excel file
            index: Whether to include the index column
            flag_values: Not used, kept for backward compatibility
            
        Returns:
            Path to the created Excel file
        """
        csv_path = Path(csv_path)
        
        # Use the CSV filename if no output name is provided
        if not output_name:
            output_name = csv_path.stem
            
        # Create output path
        output_path = self.output_dir / f"{output_name}.xlsx"
        
        try:
            # Read CSV file
            self.logger.info(f"Reading CSV file: {csv_path}")
            df = pd.read_csv(csv_path)
            self.logger.info(f"Successfully read CSV with {len(df)} rows and {len(df.columns)} columns")
            
            # Convert numeric columns to appropriate types
            self.logger.info("Converting numeric columns to appropriate types")
            df = self._convert_to_numeric(df)
            self.logger.info(f"Column types after conversion: {df.dtypes}")
            
            # Flag specific values in L1-L20 columns
            self.logger.info("Flagging values 10, 5, and 1 in L1-L20 columns")
            for col in df.columns:
                df[col] = df[col].apply(lambda x: self._flag_value(x, col))
            
            # Create Excel writer
            self.logger.info(f"Creating Excel file at: {output_path}")
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Write DataFrame to Excel
                self.logger.info(f"Writing data to sheet: {sheet_name}")
                df.to_excel(writer, sheet_name=sheet_name, index=index)
                
                # Get the worksheet
                worksheet = writer.sheets[sheet_name]
                
                # Auto-adjust column widths
                self.logger.info("Adjusting column widths")
                for idx, col in enumerate(df.columns):
                    max_length = max(
                        df[col].astype(str).apply(len).max(),
                        len(str(col))
                    )
                    worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_length + 2
            
            # Verify the file was created
            if not output_path.exists():
                raise FileNotFoundError(f"Excel file was not created at {output_path}")
            
            file_size = os.path.getsize(output_path)
            self.logger.info(f"Excel file created successfully. Size: {file_size} bytes")
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error exporting {csv_path} to Excel: {str(e)}")
            # If the file exists but is empty or corrupted, try to remove it
            if output_path.exists():
                try:
                    output_path.unlink()
                    self.logger.info(f"Removed potentially corrupted file: {output_path}")
                except Exception as cleanup_error:
                    self.logger.error(f"Failed to remove corrupted file: {cleanup_error}")
            raise

class ExcelFlagManager:
    """Utility class for flagging values in Excel files."""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def flag_values(
        self,
        excel_path: Union[str, Path],
        flag_values: List[Any],
        sheet_name: str = "Data"
    ) -> Path:
        """
        Flag specific values in an Excel file by wrapping them in !!!!.
        
        Args:
            excel_path: Path to the Excel file
            flag_values: List of values to flag
            sheet_name: Name of the sheet to process
            
        Returns:
            Path to the modified Excel file
        """
        excel_path = Path(excel_path)
        
        try:
            # Read Excel file
            self.logger.info(f"Reading Excel file: {excel_path}")
            df = pd.read_excel(excel_path, sheet_name=sheet_name)
            
            # Flag the values
            self.logger.info(f"Flagging values: {flag_values}")
            for col in df.columns:
                df[col] = df[col].apply(lambda x: f"!!!!{x}!!!!" if x in flag_values else x)
            
            # Write back to Excel
            self.logger.info("Writing flagged data back to Excel")
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get the worksheet and adjust column widths
                worksheet = writer.sheets[sheet_name]
                for idx, col in enumerate(df.columns):
                    max_length = max(
                        df[col].astype(str).apply(len).max(),
                        len(str(col))
                    )
                    worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_length + 2
            
            self.logger.info(f"Successfully flagged values in {excel_path}")
            return excel_path
            
        except Exception as e:
            self.logger.error(f"Error flagging values in {excel_path}: {str(e)}")
            raise 